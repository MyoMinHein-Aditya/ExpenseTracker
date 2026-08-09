import os
import sqlite3
from datetime import datetime
import io
from flask import Flask, request, jsonify, render_template, send_file, session, redirect, url_for
from werkzeug.security import generate_password_hash, check_password_hash
from google.oauth2 import id_token
from google.auth.transport import requests as google_requests

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", os.urandom(24).hex())  # Needed for session management
DB = "expenses.db"

# Google Auth Setup
GOOGLE_CLIENT_ID = os.environ.get("GOOGLE_CLIENT_ID", "YOUR_GOOGLE_CLIENT_ID_HERE")

class DatabaseManager:
    """Handles low-level database operations."""
    @staticmethod
    def get_conn():
        conn = sqlite3.connect(DB)
        conn.row_factory = sqlite3.Row
        return conn

    @staticmethod
    def init_db():
        with DatabaseManager.get_conn() as conn:
            # Users Table
            conn.execute("""CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                email TEXT UNIQUE,
                password TEXT,
                google_id TEXT UNIQUE
            )""")
            # Balance Table scoped by user_id
            conn.execute("""CREATE TABLE IF NOT EXISTS balance (
                user_id INTEGER PRIMARY KEY,
                amount REAL NOT NULL DEFAULT 0,
                FOREIGN KEY(user_id) REFERENCES users(id)
            )""")
            # Transactions scoped by user_id
            conn.execute("""CREATE TABLE IF NOT EXISTS transactions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                description TEXT NOT NULL,
                amount REAL NOT NULL,
                type TEXT NOT NULL,
                created_at TEXT NOT NULL,
                FOREIGN KEY(user_id) REFERENCES users(id)
            )""")
            conn.commit()

class UserManager:
    """Handles User Authentication."""
    
    @staticmethod
    def register_user(name, email, password):
        with DatabaseManager.get_conn() as conn:
            existing = conn.execute("SELECT id FROM users WHERE email=?", (email,)).fetchone()
            if existing:
                return False, "Email already registered"
                
            hashed = generate_password_hash(password)
            cur = conn.cursor()
            cur.execute("""INSERT INTO users (name, email, password) 
                           VALUES (?, ?, ?)""", (name, email, hashed))
            user_id = cur.lastrowid
            conn.execute("INSERT INTO balance (user_id, amount) VALUES (?, 0)", (user_id,))
            conn.commit()
            return True, user_id
            
    @staticmethod
    def login_user(email, password):
        with DatabaseManager.get_conn() as conn:
            user = conn.execute("SELECT id, password FROM users WHERE email=?", (email,)).fetchone()
            if not user or not user["password"]:
                return False, "Invalid email or password"
            
            if check_password_hash(user["password"], password):
                return True, user["id"]
            return False, "Invalid email or password"

    @staticmethod
    def get_or_create_google_user(google_id, email, name):
        with DatabaseManager.get_conn() as conn:
            user = conn.execute("SELECT id FROM users WHERE google_id=?", (google_id,)).fetchone()
            if user:
                return user["id"]
            
            # Check if email exists to link, or create new
            existing_email = conn.execute("SELECT id FROM users WHERE email=?", (email,)).fetchone()
            if existing_email:
                conn.execute("UPDATE users SET google_id=? WHERE id=?", (google_id, existing_email["id"]))
                conn.commit()
                return existing_email["id"]
            
            # Create New
            cur = conn.cursor()
            cur.execute("""INSERT INTO users (name, email, google_id) 
                           VALUES (?, ?, ?)""", (name, email, google_id))
            user_id = cur.lastrowid
            conn.execute("INSERT INTO balance (user_id, amount) VALUES (?, 0)", (user_id,))
            conn.commit()
            return user_id

    @staticmethod
    def verify_google_token(token):
        try:
            idinfo = id_token.verify_oauth2_token(token, google_requests.Request(), GOOGLE_CLIENT_ID)
            return {
                "google_id": idinfo['sub'],
                "email": idinfo['email'],
                "name": idinfo.get('name', 'Google User')
            }, None
        except ValueError as e:
            return None, str(e)

class ExpenseManager:
    """Handles expenses scoped securely to a specific user_id."""
    def __init__(self, user_id):
        self.user_id = user_id

    def get_balance(self):
        with DatabaseManager.get_conn() as conn:
            row = conn.execute("SELECT amount FROM balance WHERE user_id=?", (self.user_id,)).fetchone()
            return row["amount"] if row else 0.0

    def add_money(self, amount, note, created_at):
        with DatabaseManager.get_conn() as conn:
            conn.execute("UPDATE balance SET amount = amount + ? WHERE user_id=?", (amount, self.user_id))
            conn.execute(
                "INSERT INTO transactions (user_id, description, amount, type, created_at) VALUES (?, ?, ?, 'credit', ?)",
                (self.user_id, note, amount, created_at)
            )
            conn.commit()
            return self.get_balance()

    def add_expense(self, amount, description, created_at):
        with DatabaseManager.get_conn() as conn:
            bal = self.get_balance()
            if amount > bal:
                return False, "Insufficient balance"
            conn.execute("UPDATE balance SET amount = amount - ? WHERE user_id=?", (amount, self.user_id))
            conn.execute(
                "INSERT INTO transactions (user_id, description, amount, type, created_at) VALUES (?, ?, ?, 'debit', ?)",
                (self.user_id, description, amount, created_at)
            )
            conn.commit()
            return True, self.get_balance()

    def get_transactions(self, month):
        with DatabaseManager.get_conn() as conn:
            rows = conn.execute("""
                SELECT id, description, amount, type, created_at
                FROM transactions 
                WHERE user_id = ? AND strftime('%Y-%m', created_at) = ?
                ORDER BY created_at DESC""", (self.user_id, month)).fetchall()
            txns = [dict(r) for r in rows]
            total_spent = sum(t["amount"] for t in txns if t["type"] == "debit")
            total_added = sum(t["amount"] for t in txns if t["type"] == "credit")
            return txns, total_spent, total_added

    def get_months(self):
        with DatabaseManager.get_conn() as conn:
            rows = conn.execute(
                """SELECT DISTINCT strftime('%Y-%m', created_at) AS month
                FROM transactions WHERE user_id = ? ORDER BY month DESC""",
                (self.user_id,)
            ).fetchall()
        return [r["month"] for r in rows if r["month"]]

    def delete_transaction(self, txn_id):
        with DatabaseManager.get_conn() as conn:
            row = conn.execute("SELECT * FROM transactions WHERE id=? AND user_id=?", (txn_id, self.user_id)).fetchone()
            if not row:
                return False, "Not found"
            if row["type"] == "debit":
                conn.execute("UPDATE balance SET amount = amount + ? WHERE user_id=?", (row["amount"], self.user_id))
            else:
                conn.execute("UPDATE balance SET amount = amount - ? WHERE user_id=?", (row["amount"], self.user_id))
            conn.execute("DELETE FROM transactions WHERE id=?", (txn_id,))
            conn.commit()
            return True, self.get_balance()

# ── ROUTES ──────────────────────────────────────────────────────────────────

@app.before_request
def require_login():
    public_routes = ['/', '/login', '/api/auth/register', '/api/auth/login', '/api/auth/google', '/static']
    if request.path not in public_routes and not request.path.startswith('/static'):
        if 'user_id' not in session:
            if request.path.startswith('/api/'):
                return jsonify({"error": "Unauthorized"}), 401
            return redirect(url_for('login'))

@app.route("/")
def home():
    if 'user_id' in session:
        return redirect(url_for('dashboard'))
    return render_template("home.html")

@app.route("/login")
def login():
    if 'user_id' in session:
        return redirect(url_for('dashboard'))
    return render_template("login.html", google_client_id=GOOGLE_CLIENT_ID)

@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for('home'))

@app.route("/dashboard")
def dashboard():
    return render_template("index.html")

# ── AUTH API ────────────────────────────────────────────────────────────────

@app.route("/api/auth/register", methods=["POST"])
def auth_register():
    data = request.json or {}
    name = data.get("name", "").strip()
    email = data.get("email", "").strip()
    password = data.get("password", "").strip()
    
    if not name or not email or not password:
        return jsonify({"error": "All fields are required"}), 400
        
    success, result = UserManager.register_user(name, email, password)
    if not success:
        return jsonify({"error": result}), 400
        
    session['user_id'] = result
    return jsonify({"success": True})

@app.route("/api/auth/login", methods=["POST"])
def auth_login():
    data = request.json or {}
    email = data.get("email", "").strip()
    password = data.get("password", "").strip()
    
    if not email or not password:
        return jsonify({"error": "Email and password are required"}), 400
        
    success, result = UserManager.login_user(email, password)
    if not success:
        return jsonify({"error": result}), 401
        
    session['user_id'] = result
    return jsonify({"success": True})

@app.route("/api/auth/google", methods=["POST"])
def auth_google():
    data = request.json or {}
    token = data.get("token")
    if not token:
        return jsonify({"error": "Missing token"}), 400
        
    user_info, err_msg = UserManager.verify_google_token(token)
    if not user_info:
        return jsonify({"error": f"Invalid Google Token: {err_msg}"}), 400
        
    try:
        user_id = UserManager.get_or_create_google_user(
            google_id=user_info["google_id"],
            email=user_info["email"],
            name=user_info["name"]
        )
    except Exception as e:
        return jsonify({"error": f"Database Error: {str(e)}"}), 500
    
    if not user_id:
        return jsonify({"error": "Could not create user"}), 500
        
    session['user_id'] = user_id
    return jsonify({"success": True})

# ── EXPENSE API ─────────────────────────────────────────────────────────────

@app.route("/api/balance")
def get_balance():
    mgr = ExpenseManager(session['user_id'])
    return jsonify({"balance": mgr.get_balance()})

@app.route("/api/add-money", methods=["POST"])
def add_money():
    data = request.json or {}
    amount = float(data.get("amount", 0))
    note = (data.get("note") or "Added money").strip() or "Added money"

    if amount <= 0:
        return jsonify({"error": "Invalid amount"}), 400

    date_str = (data.get("date") or "").strip()
    created_at = None
    if date_str:
        try:
            created_at = datetime.strptime(date_str, "%Y-%m-%d").strftime("%Y-%m-%d %H:%M:%S")
        except ValueError:
            return jsonify({"error": "Invalid date"}), 400

    if not created_at:
        created_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    mgr = ExpenseManager(session['user_id'])
    new_bal = mgr.add_money(amount, note, created_at)
    return jsonify({"balance": new_bal})

@app.route("/api/add-expense", methods=["POST"])
def add_expense():
    data = request.json or {}
    description = (data.get("description", "") or "").strip()
    amount = float(data.get("amount", 0))
    if not description or amount <= 0:
        return jsonify({"error": "Invalid data"}), 400

    date_str = (data.get("date") or "").strip()
    created_at = None
    if date_str:
        try:
            created_at = datetime.strptime(date_str, "%Y-%m-%d").strftime("%Y-%m-%d %H:%M:%S")
        except ValueError:
            return jsonify({"error": "Invalid date"}), 400

    if not created_at:
        created_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    mgr = ExpenseManager(session['user_id'])
    success, result = mgr.add_expense(amount, description, created_at)
    if not success:
        return jsonify({"error": result}), 400
    return jsonify({"balance": result})

@app.route("/api/transactions")
def get_transactions():
    month = request.args.get("month", datetime.now().strftime("%Y-%m"))
    mgr = ExpenseManager(session['user_id'])
    txns, total_spent, total_added = mgr.get_transactions(month)
    return jsonify({"transactions": txns, "total_spent": total_spent, "total_added": total_added})

@app.route("/api/months")
def get_months():
    mgr = ExpenseManager(session['user_id'])
    months = mgr.get_months()
    if not months:
        months = [datetime.now().strftime("%Y-%m")]
    return jsonify({"months": months})

@app.route("/api/delete/<int:txn_id>", methods=["DELETE"])
def delete_transaction(txn_id):
    mgr = ExpenseManager(session['user_id'])
    success, result = mgr.delete_transaction(txn_id)
    if not success:
        return jsonify({"error": result}), 404
    return jsonify({"balance": result})

# ── CHARTS ──────────────────────────────────────────────────────────────────

def chart_style(fig, ax):
    BG = "#110E17"
    fig.patch.set_facecolor(BG)
    ax.set_facecolor("#110E17")
    ax.tick_params(colors="#2D2838", labelsize=8)
    ax.spines[:].set_color("#2D2838")
    for spine in ax.spines.values():
        spine.set_linewidth(0.5)

@app.route("/api/chart/daily")
def chart_daily():
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    import calendar

    month = request.args.get("month", datetime.now().strftime("%Y-%m"))
    year, mon = map(int, month.split("-"))
    days_in_month = calendar.monthrange(year, mon)[1]

    with DatabaseManager.get_conn() as conn:
        rows = conn.execute("""
            SELECT CAST(strftime('%d', created_at) AS INTEGER) as day,
                   SUM(CASE WHEN type='debit' THEN amount ELSE 0 END) as spent
            FROM transactions
            WHERE user_id=? AND strftime('%Y-%m', created_at) = ?
            GROUP BY day ORDER BY day
        """, (session['user_id'], month)).fetchall()

    data = {r["day"]: r["spent"] for r in rows}
    days = list(range(1, days_in_month + 1))
    values = [data.get(d, 0) for d in days]

    fig, ax = plt.subplots(figsize=(7, 2.6))
    chart_style(fig, ax)

    colors = ["#FFD6A5" if v == max(values) and v > 0 else "#6A00F4" if v > 0 else "#E5E5E6" for v in values]
    ax.bar(days, values, color=colors, width=0.7, zorder=3)
    ax.set_xlim(0.5, days_in_month + 0.5)
    ax.set_xlabel("Day of Month", color="#F7F3EE", fontsize=8, labelpad=6)
    ax.set_ylabel("₹ Spent", color="#F7F3EE", fontsize=8, labelpad=6)
    ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, _: f"₹{x/1000:.0f}k" if x >= 1000 else f"₹{x:.0f}"))
    ax.grid(axis="y", color="#c98b47", linewidth=0.5, zorder=0)
    ax.set_title("Daily Spending — " + datetime.strptime(month, "%Y-%m").strftime("%B %Y"),
                 color="#F7F3EE", fontsize=10, pad=10, fontweight="bold")
    plt.tight_layout(pad=1.2)

    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=130, bbox_inches="tight", facecolor=fig.get_facecolor())
    buf.seek(0)
    plt.close(fig)
    return send_file(buf, mimetype="image/png")

@app.route("/api/chart/monthly")
def chart_monthly():
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    import numpy as np

    with DatabaseManager.get_conn() as conn:
        rows = conn.execute("""
            SELECT strftime('%Y-%m', created_at) as month, type, SUM(amount) as total
            FROM transactions WHERE user_id=? GROUP BY month, type ORDER BY month ASC
        """, (session['user_id'],)).fetchall()

    months_data = {}
    for r in rows:
        m = r["month"]
        months_data.setdefault(m, {"spent": 0, "added": 0})
        if r["type"] == "debit":
            months_data[m]["spent"] = r["total"]
        else:
            months_data[m]["added"] = r["total"]

    sorted_months = sorted(months_data.keys())[-6:]
    labels = [datetime.strptime(m, "%Y-%m").strftime("%b '%y") for m in sorted_months]
    spent  = [months_data[m]["spent"]  for m in sorted_months]
    added  = [months_data[m]["added"]  for m in sorted_months]

    x = np.arange(len(labels))
    w = 0.35

    fig, ax = plt.subplots(figsize=(7, 2.6))
    chart_style(fig, ax)

    ax.bar(x - w/2, added, w, label="Income",  color="#6A00F4", alpha=0.9, zorder=3)
    ax.bar(x + w/2, spent, w, label="Expenses", color="#FFD6A5", alpha=0.9, zorder=3)
    ax.set_xticks(x)
    ax.set_xticklabels(labels, color="#F7F3EE", fontsize=8)
    ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda v, _: f"₹{v/1000:.0f}k" if v >= 1000 else f"₹{v:.0f}"))
    ax.grid(axis="y", color="#2D2838", linewidth=0.5, zorder=0)
    ax.legend(fontsize=8, facecolor="#110E17", edgecolor="#2D2838",
              labelcolor="#F7F3EE", loc="upper left")
    ax.set_title("Monthly Overview (Last 6 Months)", color="#F7F3EE", fontsize=10, pad=10, fontweight="bold")
    plt.tight_layout(pad=1.2)

    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=130, bbox_inches="tight", facecolor=fig.get_facecolor())
    buf.seek(0)
    plt.close(fig)
    return send_file(buf, mimetype="image/png")

@app.route("/api/chart/breakdown")
def chart_breakdown():
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt

    month = request.args.get("month", datetime.now().strftime("%Y-%m"))
    with DatabaseManager.get_conn() as conn:
        rows = conn.execute("""
            SELECT description, SUM(amount) as total
            FROM transactions
            WHERE user_id=? AND type='debit' AND strftime('%Y-%m', created_at) = ?
            GROUP BY description ORDER BY total DESC LIMIT 7
        """, (session['user_id'], month)).fetchall()

    if not rows:
        fig, ax = plt.subplots(figsize=(4, 2.6))
        chart_style(fig, ax)
        ax.text(0.5, 0.5, "No expenses yet", ha="center", va="center",
                color="#F7F3EE", fontsize=10, transform=ax.transAxes)
        ax.axis("off")
        ax.set_title("Expense Breakdown", color="#F7F3EE", fontsize=10, pad=10, fontweight="bold")
        buf = io.BytesIO()
        fig.savefig(buf, format="png", dpi=130, bbox_inches="tight", facecolor=fig.get_facecolor())
        buf.seek(0); plt.close(fig)
        return send_file(buf, mimetype="image/png")

    labels = [r["description"] for r in rows]
    values = [r["total"] for r in rows]
    palette = ["#6A00F4", "#FFD6A5", "#00A389", "#D0B3FF", "#E59560"]

    fig, ax = plt.subplots(figsize=(4, 2.6))
    chart_style(fig, ax)
    wedges, texts, autotexts = ax.pie(
        values, labels=None, autopct="%1.0f%%",
        colors=palette[:len(values)], startangle=90,
        wedgeprops=dict(width=0.55, edgecolor="#110E17", linewidth=2),
        pctdistance=0.75
    )
    for at in autotexts:
        at.set_color("#FFFFFF"); at.set_fontsize(7); at.set_fontweight("bold")

    ax.legend(wedges, [f"{l[:14]}" for l in labels],
              fontsize=7, facecolor="#110E17", edgecolor="#2D2838",
              labelcolor="#F7F3EE", loc="center left",
              bbox_to_anchor=(0.85, 0.5), framealpha=0.9)
    ax.set_title("Expense Breakdown", color="#F7F3EE", fontsize=10, pad=10, fontweight="bold")
    plt.tight_layout(pad=0.5)

    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=130, bbox_inches="tight", facecolor=fig.get_facecolor())
    buf.seek(0); plt.close(fig)
    return send_file(buf, mimetype="image/png")

# ── EXPORT ──────────────────────────────────────────────────────────────────

@app.route("/api/export")
def export_excel():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return jsonify({"error": "Run: pip install openpyxl"}), 500

    month = request.args.get("month", datetime.now().strftime("%Y-%m"))
    mgr = ExpenseManager(session['user_id'])
    
    with DatabaseManager.get_conn() as conn:
        rows = conn.execute("""SELECT description, amount, type, created_at
            FROM transactions WHERE user_id=? AND strftime('%Y-%m', created_at) = ?
            ORDER BY created_at ASC""", (session['user_id'], month)).fetchall()
        txns = [dict(r) for r in rows]
        bal = mgr.get_balance()

    wb = openpyxl.Workbook()
    ws = wb.active
    month_label = datetime.strptime(month, "%Y-%m").strftime("%B %Y")
    ws.title = month_label

    dark  = PatternFill("solid", fgColor="0E0E0E")
    dark2 = PatternFill("solid", fgColor="1A1A1A")
    rowa  = PatternFill("solid", fgColor="111111")
    rowb  = PatternFill("solid", fgColor="161616")
    facc  = Font(name="Calibri", bold=True, color="C8F04A", size=14)
    fhdr  = Font(name="Calibri", bold=True, color="FFFFFF",  size=11)
    fred  = Font(name="Calibri", bold=True, color="FF5F5F",  size=11)
    fgrn  = Font(name="Calibri", bold=True, color="4AFFA0",  size=11)
    fyel  = Font(name="Calibri", bold=True, color="C8F04A",  size=11)
    fmut  = Font(name="Calibri", color="888888", size=10)
    fnor  = Font(name="Calibri", color="F0F0F0", size=11)
    thin  = Border(bottom=Side(style="thin",   color="2E2E2E"))
    thick = Border(bottom=Side(style="medium", color="C8F04A"))
    ctr   = Alignment(horizontal="left", vertical="center")

    def sc(cell, val, font, fill, align=ctr):
        cell.value=val; cell.font=font; cell.fill=fill; cell.alignment=align

    total_spent = sum(t["amount"] for t in txns if t["type"] == "debit")
    total_added = sum(t["amount"] for t in txns if t["type"] == "credit")

    ws.merge_cells("A1:D1"); sc(ws["A1"], f"PaisaTrack — {month_label}", facc, dark); ws.row_dimensions[1].height=32
    ws.merge_cells("A2:B2"); sc(ws["A2"], f"Total Added:  ₹{total_added:,.2f}", fgrn, dark)
    ws.merge_cells("C2:D2"); sc(ws["C2"], f"Total Spent:  ₹{total_spent:,.2f}", fred, dark); ws.row_dimensions[2].height=22
    ws.merge_cells("A3:D3"); sc(ws["A3"], f"Current Balance:  ₹{bal:,.2f}", fyel, dark); ws.row_dimensions[3].height=22
    for c in "ABCD": ws[f"{c}4"].fill=dark
    ws.row_dimensions[4].height=8

    for i,h in enumerate(["Date & Time","Description","Type","Amount (₹)"],1):
        cell=ws.cell(row=5,column=i,value=h)
        cell.font=fhdr; cell.fill=dark2; cell.alignment=ctr; cell.border=thick
    ws.row_dimensions[5].height=22

    for ri, t in enumerate(txns, 6):
        dt = datetime.strptime(t["created_at"], "%Y-%m-%d %H:%M:%S")
        debit = t["type"]=="debit"
        fill = rowa if ri%2==0 else rowb
        vals = [dt.strftime("%d %b, %I:%M %p"), t["description"],
                "Expense" if debit else "Added",
                f"{'−' if debit else '+'} ₹{t['amount']:,.2f}"]
        fonts = [fmut, fnor, Font(name="Calibri",color="FF5F5F" if debit else "4AFFA0",size=10),
                 fred if debit else fgrn]
        for ci,(v,f) in enumerate(zip(vals,fonts),1):
            cell=ws.cell(row=ri,column=ci,value=v)
            cell.font=f; cell.fill=fill; cell.border=thin; cell.alignment=ctr
        ws.row_dimensions[ri].height=20

    ws.column_dimensions["A"].width=22
    ws.column_dimensions["B"].width=36
    ws.column_dimensions["C"].width=12
    ws.column_dimensions["D"].width=20
    ws.sheet_view.showGridLines=False

    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name=f"PaisaTrack_{month}.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

if __name__ == "__main__":
    DatabaseManager.init_db()
    print("\n PaisaTrack Auth version running at http://localhost:5000\n")
    app.run(debug=True, port=5000)
